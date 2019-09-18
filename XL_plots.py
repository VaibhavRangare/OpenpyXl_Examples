import openpyxl as xl
from openpyxl.chart import BarChart, Reference, BarChart3D, PieChart, PieChart3D, BubbleChart, LineChart, LineChart3D, \
    StockChart, ScatterChart, SurfaceChart, AreaChart, AreaChart3D


def plot_area_chart3_d():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    seriesobject = xl.chart.Series(values, title="Yearly Revenue")
    chart = AreaChart3D()
    chart.append(seriesobject)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


def plot_area_chart():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    series_object = xl.chart.Series(values, title="Yearly Revenue")
    chart = AreaChart()
    chart.append(series_object)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


def plot_surface_chart():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    series_object = xl.chart.Series(values, title="Yearly Revenue")
    chart = SurfaceChart()
    chart.append(series_object)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


def plot_scatter_chart():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    series_object = xl.chart.Series(values, title="Yearly Revenue")
    chart = ScatterChart()
    chart.append(series_object)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


def plot_stock_chart():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    series_object = xl.chart.Series(values, title="Yearly Revenue")
    chart = StockChart()
    chart.append(series_object)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


def plot_line_chart3_d():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    series_object = xl.chart.Series(values, title="Yearly Revenue")
    chart = LineChart3D()
    chart.append(series_object)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


def plot_line_chart():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    series_object = xl.chart.Series(values, title="Yearly Revenue")
    chart = LineChart()
    chart.append(series_object)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


def plot_bubble_chart():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    series_object = xl.chart.Series(values, title="Yearly Revenue")
    chart = BubbleChart()
    chart.append(series_object)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


def plot_bar_chart():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    series_object = xl.chart.Series(values, title="Yearly Revenue")
    chart = BarChart()
    chart.append(series_object)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


def plot_bar_chart3_d():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    series_object = xl.chart.Series(values, title="Yearly Revenue")
    chart = BarChart3D()
    chart.append(series_object)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


def plot_pie_chart():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    series_object = xl.chart.Series(values, title="Yearly Revenue")
    chart = PieChart()
    chart.append(series_object)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


def plot_pie_chart3_d():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')  # load xl workbook
    sheet = wb['dATA']  # get the worksheet
    for i in range(2, 10):
        cell = sheet['f'+str(i)]
        sheet['a'+str(i)] = cell.value
    values = Reference(
        sheet,
        min_row=2,
        max_row=10,
        min_col=1,
        max_col=1
    )
    series_object = xl.chart.Series(values, title="Yearly Revenue")
    chart = PieChart3D()
    chart.append(series_object)
    sheet.add_chart(chart, 'j2')
    wb.save('C:\\Resources\\Program_Xl_Sheets\\Groceriesptfinal.xlsx')


plot_area_chart()
plot_surface_chart()            # Not working
plot_scatter_chart()            # Not working
plot_stock_chart()
plot_line_chart3_d()
plot_line_chart()
plot_bubble_chart()             # Not working
plot_bar_chart()
plot_bar_chart3_d()
plot_pie_chart()
plot_pie_chart3_d()
