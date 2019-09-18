from openpyxl import Workbook


# Create an xl workbook with one sheet
# Workbook comes with one default sheet named 'Sheet', position of this sheet is 0
wb = Workbook()
print("sheet Names:")
print(wb.sheetnames)                                                # Gives Default Sheet
sheet = wb.active                                                   # Get Sheet
# sheet = wb['sheet_name']                                          # get Sheet
wb.create_sheet("First_Sheet", 0)                                    # Create a first sheet
wb.create_sheet("End_Sheet")                                        # Last Sheet
# sheet['cell_name'] = value
sheet['A1'] = "Name"
print(wb.sheetnames)                                                # Provides all sheet names
# wb.save('SampleWorkBook.xlsx')                                     # Save workbook at the project location
wb.save('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')    # Save workbook at specific directory
