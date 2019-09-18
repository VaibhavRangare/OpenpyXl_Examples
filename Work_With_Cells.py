import openpyxl as xl


def change_cell_value():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')
    sheet = wb['First_Sheet']                                           # Get a sheet from workbook
    cell = sheet['a1']                                                  # get a cell from sheet
    cell.value = "College"                                              # change the value of cell
    cell = sheet.cell(1, 2)                                              # get a cell from sheet
    cell.value = "Price"                                           # change the value of cell
    wb.save('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')       # Save workbook


def add_multiple_values():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')
    sheet = wb['First_Sheet']  # Get a sheet from workbook
    for i in range(2, 11):
        cell = sheet.cell(i, 2)
        cell.value = i
        i += i
    wb.save('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')  # Save workbook


def delete_multiple_values():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')
    sheet = wb['First_Sheet']  # Get a sheet from workbook
    for i in range(2, 11):
        cell = sheet.cell(i, 1)
        cell.value = ""
        i += i
    wb.save('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')  # Save workbook


def delete_value():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')
    sheet = wb['First_Sheet']  # Get a sheet from workbook
    cell = sheet.cell(1, 1)
    cell.value = ""
    wb.save('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')  # Save workbook


def add_value():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')
    sheet = wb['First_Sheet']  # Get a sheet from workbook
    cell = sheet.cell(11, 1)
    cell.value = "SUM"
    wb.save('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')  # Save workbook


def apply_rule_to_column():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')
    sheet = wb['First_Sheet']  # Get a sheet from workbook
    sum_value = 0
    for i in range(2, 11):
        sum_value += i
    cell = sheet.cell(11, 2)
    cell.value = sum_value
    wb.save('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')  # Save workbook


def modify_row_cells():
    wb = xl.load_workbook('C:\\Resources\\Program_Xl_Sheets\\SampleWorkBook.xlsx')
    sheet = wb['First_Sheet']                                           # Get a sheet from workbook
    print(sheet.max_row)
    print(sheet.max_column)


# Change_CellValue()
# delete_Multiple_Values()
# delete_Value()
# modify_row_cells()
# add_Multiple_values()
# add_Value()
apply_rule_to_column()
